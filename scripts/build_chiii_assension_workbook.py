#!/usr/bin/env python3
"""Build the Assension dissertation-event workbook from templates + the uploaded artifact offering."""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DB = ROOT / "_CHPT III_Assension_db"
TPL = DB / "templates"
SRC = DB / "uploads" / "ITDRPaaS_Governance_Artifact_Workbook.xlsx"
OUT = DB / "uploads" / "ITDRPaaS_Governance_Artifact_Workbook_Assension.xlsx"

HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="1F4D6D")
TITLE = Font(bold=True, size=14, color="1F4D6D")
WRAP = Alignment(wrap_text=True, vertical="top")


def header(ws, titles, row=1):
    for col, title in enumerate(titles, 1):
        cell = ws.cell(row, col, title)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = WRAP


def autosize(ws, max_width=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max((len(str(c.value or "")) for c in col), default=12), max_width)
        ws.column_dimensions[letter].width = max(width, 14)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP


def read_csv(name: str) -> tuple[list[str], list[dict]]:
    path = TPL / name
    with path.open(encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))
        fields = list(rows[0].keys()) if rows else []
    return fields, rows


def add_csv_sheet(wb, title, csv_name, blurb):
    fields, rows = read_csv(csv_name)
    ws = wb.create_sheet(title)
    ws["A1"] = blurb
    ws["A1"].font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(fields), 3))
    header(ws, fields, row=3)
    for r_i, rec in enumerate(rows, 4):
        for c_i, key in enumerate(fields, 1):
            ws.cell(r_i, c_i, rec.get(key, ""))
    autosize(ws)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(fields))}{3 + len(rows)}"
    return ws


def copy_artifact_sheets(wb):
    src = load_workbook(SRC, data_only=False)
    keep = [
        "Master Register",
        "Responsibility Matrix",
        "DR Plan",
        "Incident Command",
        "Delegation Authority",
        "Failover Authorization",
        "Risk Exception",
        "Audit Trail",
        "Lists",
    ]
    for name in keep:
        if name not in src.sheetnames:
            continue
        src_ws = src[name]
        dst = wb.create_sheet(name[:31])
        for row in src_ws.iter_rows():
            for cell in row:
                dst.cell(cell.row, cell.column, cell.value)
        # Improve: Chapter / dissertation-event stamp in column A if empty-ish
        if dst["A1"].value in (None, ""):
            dst["A1"] = "Chapter III secondary instrument (artifacts as context). Dissertation event: Methodology / Measures."
        dst.freeze_panes = src_ws.freeze_panes
    src.close()


def main() -> None:
    wb = Workbook()
    readme = wb.active
    readme.title = "00_ReadMe"
    header(readme, ["Section", "Text"])
    rows = [
        ("Purpose", "Assension twin of the ITDRPaaS Governance Artifact Workbook. Adds dissertation-event tabs (Chapter I–V) and specialty tabs (Theory, Qualitative, U.S. SME, Qualitative Analysis, Lit review) without dropping the operational artifact sheets."),
        ("Spine", "SAL → DR → ESC → AE → RA is an organizing conceptual model, not a substantive theory."),
        ("PLU rule", "Power = capacity to influence or authorize. Legitimacy = recognized authority or appropriate standing. Urgency = time sensitivity. Combinations allowed. Leverage is the enacted combination, not a fourth attribute."),
        ("Chapter split", "Chapter III plans. Chapter IV reports. Chapter V interprets. Theme/Finding cells stay empty here."),
        ("Confidentiality", "No credentials, IP addresses, customer identifiers, or unredacted diagrams. Use P##, role, SME-A, Platform-1."),
        ("Source of truth", "_CHPT III_Assension_db/templates/*.csv and registers/*.csv. Rebuild with python3 scripts/build_chiii_assension_workbook.py."),
        ("Uploaded offering", "Original operational workbook retained as ITDRPaaS_Governance_Artifact_Workbook.xlsx in the same uploads folder."),
    ]
    for r in rows:
        readme.append(list(r))
    autosize(readme, 80)

    add_csv_sheet(wb, "01_ChI_Introduction", "chapter-i-needed-items.csv", "Chapter I dissertation event: items needed inside the introduction.")
    add_csv_sheet(wb, "02_ChII_LitReview", "chapter-ii-lit-review.csv", "Chapter II dissertation event: literature-review needed items.")
    add_csv_sheet(wb, "03_ChIII_Methodology", "chapter-iii-needed-items.csv", "Chapter III dissertation event: every methodology header with a how-to link.")
    add_csv_sheet(wb, "04_ChIV_Findings", "chapter-iv-needed-items.csv", "Chapter IV dissertation event: reserved findings items. Theme titles stay empty until meaning units are locked.")
    add_csv_sheet(wb, "05_ChV_Discussion", "chapter-v-needed-items.csv", "Chapter V dissertation event: interpretation only. Do not newly name themes.")
    add_csv_sheet(wb, "06_Theory", "theory.csv", "Theory tab: PLU, leverage, DR/ESC/AE/RA, proposed process model.")
    add_csv_sheet(wb, "07_Qualitative", "qualitative.csv", "Qualitative design tab: GQI, CIT, instrument, artifacts as context.")
    add_csv_sheet(wb, "08_US_SME", "us-sme.csv", "U.S. SME population, inclusion, sample, and recruitment rules.")
    add_csv_sheet(wb, "09_Qual_Analysis", "qualitative-analysis.csv", "Qualitative analysis tab: hybrid codebook TA, Delve as CAQDAS only, control rules.")
    add_csv_sheet(wb, "10_Lit_Review", "lit-review.csv", "Lit-review use rules: original vs build-on vs new instrument vs audit artifact.")

    copy_artifact_sheets(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT} sheets={wb.sheetnames}")


if __name__ == "__main__":
    main()
