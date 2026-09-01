#!/usr/bin/env python3
"""Build the amalgamated Excel workbook with the same tab order as the site."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUT = ROOT / "workbook" / "BMGT8044_Amalgamated_Research_Workspace.xlsx"

HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="1F4D6D")
WRAP = Alignment(wrap_text=True, vertical="top")


def header(ws, titles):
    ws.append(titles)
    for cell in ws[1]:
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = WRAP


def autosize(ws, max_width=48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max((len(str(c.value or "")) for c in col), default=12), max_width)
        ws.column_dimensions[letter].width = max(width, 12)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP


def load(name):
    return json.loads((DATA / name).read_text(encoding="utf-8"))


def main() -> None:
    counts = load("counts.json")
    wb = Workbook()

    readme = wb.active
    readme.title = "00_README"
    header(readme, ["Section", "Text"])
    rows = [
        ("Purpose", "Amalgamated BMGT 8044 workspace for generic qualitative interviews. No source fields were dropped."),
        ("Tab order", "README, Dashboard (search + last three), Parking lot, References, Theory & spine, Leadership, Qual methods, Source archive, RQ1, Interview protocol, Analysis templates, Chapter III Assension."),
        ("Live vs archive", "Quantitative classmate Qualtrics/regression material is ARCHIVE_COMPARATIVE only."),
        ("IDs", "REF-, XREF-, THEME-, PL-, PL-L, TH-, SP-, IQ-, ARC-, CH3-H-, CH3-AT-, CH3-SP-"),
        ("Search rule", "Dashboard is the only search surface. Last three distinct searches persist in the browser and are mirrored in LastSearch1-3."),
        ("RQ1", "Evaluate the appropriateness of data analysis strategies for qualitative methodologies. Reflexive thematic analysis is the primary appropriate strategy for this interview study."),
    ]
    for r in rows:
        readme.append(list(r))
    autosize(readme, 80)

    dash = wb.create_sheet("01_DASHBOARD")
    header(dash, ["Metric", "Value"])
    for k, v in counts.items():
        dash.append([k, v])
    dash.append([])
    dash.append(["LastSearch1", ""])
    dash.append(["LastSearch2", ""])
    dash.append(["LastSearch3", ""])
    dash.append(["Note", "Browser localStorage is the live last-three store. Export search_history.json from the Dashboard to paste here."])
    autosize(dash, 70)

    pl = wb.create_sheet("02_PARKING_LOT")
    header(pl, ["PL-ID", "SourceTab", "LinkedID", "Item", "WhyParked", "Status", "Owner", "Due", "Resolution", "ReturnedToTab"])
    for rec in load("parking_lot.json"):
        pl.append([
            rec.get("id"),
            rec.get("source_tab"),
            rec.get("linked_id"),
            rec.get("title"),
            rec.get("why_parked"),
            rec.get("status"),
            rec.get("owner"),
            rec.get("due"),
            rec.get("resolution"),
            rec.get("returned_to_tab"),
        ])
    autosize(pl)

    refs = wb.create_sheet("03_REFERENCES_MASTER")
    header(refs, ["ID", "Week/section", "Type", "Title", "Author", "Journal", "Year", "DOI", "URL", "Permalink", "Student note", "UsedInTab"])
    for rec in load("references.json"):
        refs.append([
            rec.get("id"),
            rec.get("section_name"),
            rec.get("item_type"),
            rec.get("item_title"),
            rec.get("item_author"),
            rec.get("item_journal_title"),
            rec.get("item_publication_date"),
            rec.get("item_doi"),
            rec.get("item_url"),
            rec.get("permalink"),
            rec.get("item_student_note"),
            rec.get("used_in_tab"),
        ])
    for rec in load("extra_refs.json"):
        refs.append([
            rec.get("id"),
            rec.get("week"),
            rec.get("item_type"),
            rec.get("title"),
            "",
            "",
            "",
            "",
            rec.get("item_url"),
            "",
            rec.get("apa"),
            rec.get("used_in_tab"),
        ])
    autosize(refs)

    theory = wb.create_sheet("04_THEORY_AND_SPINE")
    header(theory, ["ID", "Type", "Title", "Text", "Source tab", "Source row"])
    for rec in load("spine.json"):
        theory.append([rec.get("id"), "Spine", rec.get("title"), rec.get("description"), rec.get("source_tab"), rec.get("source_row")])
    for rec in load("theory.json"):
        theory.append([rec.get("id"), "Theory", rec.get("title"), rec.get("description"), rec.get("source_tab"), rec.get("source_row")])
    autosize(theory)

    lead = wb.create_sheet("05_LEADERSHIP_ALIGNMENT")
    header(lead, ["ID", "Theme", "Description", "References", "Source file"])
    for rec in load("themes.json"):
        lead.append([rec.get("id"), rec.get("theme"), rec.get("description"), rec.get("references"), rec.get("source_file")])
    autosize(lead)

    methods = wb.create_sheet("06_QUAL_METHODS_AND_TOOLS")
    header(methods, ["ID", "Source", "Title", "Text"])
    for rec in load("archive.json"):
        if rec.get("used_in_tab") == "06_QUAL_METHODS_AND_TOOLS":
            methods.append([rec.get("id"), rec.get("item_type"), rec.get("title"), rec.get("description")])
    autosize(methods)

    archive = wb.create_sheet("07_SOURCE_ARCHIVE")
    header(archive, ["ID", "Class", "Source file", "Source tab", "Title", "Full text"])
    for rec in load("archive.json"):
        if rec.get("used_in_tab") != "06_QUAL_METHODS_AND_TOOLS":
            archive.append([rec.get("id"), rec.get("item_type"), rec.get("source_file"), rec.get("source_tab"), rec.get("title"), rec.get("description")])
    autosize(archive)

    rq1 = wb.create_sheet("08_RQ1_ANALYSIS")
    header(rq1, ["Section", "Third-person statement"])
    rq1.append(["Research question", "Evaluate the appropriateness of data analysis strategies for use with qualitative research methodologies."])
    rq1.append(["Design", "Generic qualitative inquiry with semi-structured critical-incident interviews. Stakeholder salience (power, legitimacy, urgency) is the operational lens."])
    rq1.append(["Primary strategy", "Reflexive thematic analysis is appropriate because it stays with participant meaning and still allows a declared lens to sensitize coding."])
    rq1.append(["Companion", "Theory-informed thematic analysis is appropriate as a companion. CAQDAS/NVivo is an audit environment, not a method."])
    rq1.append(["Not appropriate", "Survey scoring, simple linear regression, and a priori power analysis as substitutes for information power belong to the archived comparative sheet."])
    rq1.append(["Ethical control", "A review question must test the same construct as its parent. Unaligned prompts must not be used for extraction."])
    rq1.append(["What alignment does", "Aligned tables create an audit trail from citation to construct to question to excerpt to code to theme to parked exception."])
    autosize(rq1, 80)

    iq = wb.create_sheet("09_INTERVIEW_PROTOCOL")
    header(iq, ["IQ-ID", "Construct", "Interview question", "Review question", "Ethical extraction check", "Linked IDs", "Status"])
    for rec in load("interview_questions.json"):
        linked = rec.get("linked_refs")
        if isinstance(linked, list):
            linked = ", ".join(linked)
        iq.append([
            rec.get("id"),
            rec.get("construct"),
            rec.get("question"),
            rec.get("review_question"),
            rec.get("ethical_check"),
            linked,
            rec.get("status"),
        ])
    autosize(iq, 60)

    tmpl = wb.create_sheet("10_ANALYSIS_TEMPLATES")
    header(tmpl, ["Template", "Col1", "Col2", "Col3", "Col4", "Col5", "Col6"])
    tmpl.append(["Interview log", "Interview ID", "Date", "Role", "IQ IDs", "Member-check", "Notes"])
    tmpl.append(["Codebook", "Code ID", "Name", "Definition", "Inclusion", "Exclusion", "Linked construct"])
    tmpl.append(["Excerpt-to-code", "Excerpt ID", "Interview ID", "IQ-ID", "Verbatim", "Code IDs", "Why eligible"])
    tmpl.append(["Theme development", "Theme ID", "Candidate", "Support excerpts", "Negative excerpts", "Construct", "Decision"])
    tmpl.append(["Negative case", "Case ID", "Predicted", "Observed", "Revision", "Status", ""])
    tmpl.append(["Memo / audit", "Memo ID", "Date", "Type", "Linked IDs", "Memo", "Decision"])
    tmpl.append(["Parking intake", "PL-ID", "Source tab", "Linked ID", "Item", "Why parked", "Resolution"])
    autosize(tmpl, 28)

    lineage = wb.create_sheet("11_SOURCE_LINEAGE")
    header(lineage, ["Source file", "Source tab", "Nonempty cells", "Destination", "Notes"])
    for rec in json.loads((ROOT / "source" / "inventory" / "lineage.json").read_text(encoding="utf-8")):
        lineage.append([rec.get("source_file"), rec.get("source_tab"), rec.get("nonempty_cells"), rec.get("destination_tab"), rec.get("notes")])
    autosize(lineage, 50)

    chiii = wb.create_sheet("12_CHIII_ASSENSION")
    header(chiii, ["ID", "Level", "Heading", "MD link", "How-to", "Spine nodes", "Alignment testers", "Status"])
    nest = json.loads((DATA / "chiii_assension.json").read_text(encoding="utf-8"))
    for rec in nest.get("headers", []):
        chiii.append([
            rec.get("id"),
            rec.get("level"),
            rec.get("heading"),
            rec.get("md_link"),
            rec.get("howto_link"),
            rec.get("spine_nodes"),
            rec.get("alignment_nodes"),
            rec.get("tester_status"),
        ])
    autosize(chiii, 48)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
