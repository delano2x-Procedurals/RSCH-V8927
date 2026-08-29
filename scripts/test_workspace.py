#!/usr/bin/env python3
"""Sanity checks for extracted IDs, pages, and last-three search rules."""
from __future__ import annotations

import json
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]


def last_three(existing, query, facets):
    entry = {"q": query.strip(), "facets": facets}
    if not entry["q"] and not any(facets.values()):
        return existing[:3]
    nxt = [e for e in existing if not (e["q"] == entry["q"] and e["facets"] == entry["facets"])]
    nxt.insert(0, entry)
    return nxt[:3]


def main() -> None:
    counts = json.loads((ROOT / "data" / "counts.json").read_text())
    assert counts["references"] == 31, counts
    assert counts["themes"] == 9, counts
    assert counts["interview_aligned"] == 8
    assert counts["interview_unaligned"] == 1
    assert counts["index"] > 1000

    refs = json.loads((ROOT / "data" / "references.json").read_text())
    assert refs[0]["id"] == "REF-001"
    assert any("NVivo" in (r.get("item_title") or "") for r in refs)

    themes = json.loads((ROOT / "data" / "themes.json").read_text())
    assert any(t["theme"] == "Servant Leadership" for t in themes)

    lineage = json.loads((ROOT / "source" / "inventory" / "lineage.json").read_text())
    assert len(lineage) >= 20

    pages = [
        "index.html",
        "dashboard.html",
        "parking-lot.html",
        "references.html",
        "theory-spine.html",
        "leadership.html",
        "methods.html",
        "archive.html",
        "rq1.html",
        "interview-protocol.html",
        "analysis-templates.html",
        "citations.html",
        "sop.html",
    ]
    for name in pages:
        text = (ROOT / "workspace-app" / name).read_text()
        assert 'class="toc"' in (ROOT / "workspace-app" / "js" / "nav.js").read_text()
        assert "<h2" in text

    dash = (ROOT / "workspace-app" / "dashboard.html").read_text()
    assert "last-three" in dash
    assert "Expedited filter" in dash

    readme = (ROOT / "workspace-app" / "index.html").read_text()
    assert "README" in readme

    rq1 = (ROOT / "workspace-app" / "rq1.html").read_text()
    assert "The member" in rq1
    assert "I will" not in rq1
    assert "Reflexive thematic analysis" in rq1
    assert "Delve" in rq1
    methods = (ROOT / "workspace-app" / "methods.html").read_text()
    assert "Delve" in methods
    assert "CAQDAS product" in methods
    cites = (ROOT / "workspace-app" / "citations.html").read_text()
    assert "Parenthetical insert" in cites
    sop = (ROOT / "workspace-app" / "sop.html").read_text()
    assert "RQ1 data statement" in sop
    assert "(Lester et al., 2020)" in sop
    assert "Delve (n.d.)" in sop
    assert "I will" not in sop
    assert (ROOT / "docs" / "RQ1_SOP_Data_Statement.md").exists()
    nav = (ROOT / "workspace-app" / "js" / "nav.js").read_text()
    assert "Save citation insert" in nav
    assert "citations.html" in nav

    index = json.loads((ROOT / "data" / "index.json").read_text())
    assert any(r.get("id") == "TOOL-DELVE" for r in index)

    hist = []
    hist = last_three(hist, "NVivo", {"type": "", "week": "", "status": "", "used": ""})
    hist = last_three(hist, "servant", {"type": "Theme", "week": "", "status": "", "used": ""})
    hist = last_three(hist, "salience", {"type": "", "week": "", "status": "", "used": ""})
    hist = last_three(hist, "ethics", {"type": "", "week": "", "status": "", "used": ""})
    assert len(hist) == 3
    assert hist[0]["q"] == "ethics"
    assert hist[1]["q"] == "salience"
    assert hist[2]["q"] == "servant"
    hist = last_three(hist, "ethics", {"type": "", "week": "", "status": "", "used": ""})
    assert hist[0]["q"] == "ethics"
    assert len(hist) == 3

    xlsx = ROOT / "workbook" / "BMGT8044_Amalgamated_Research_Workspace.xlsx"
    assert xlsx.exists()
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, read_only=True)
    names = wb.sheetnames
    assert "11_CITATION_INSERTS" in names
    assert "12_TOOL_USAGE" in names
    assert "14_SOP_RQ1" in names
    assert "00_README" in names
    cite = wb["11_CITATION_INSERTS"]
    headers = [c.value for c in next(cite.iter_rows(min_row=1, max_row=1))]
    assert "ParentheticalInsert" in headers
    wb.close()
    print("test_workspace: ok")


if __name__ == "__main__":
    main()
